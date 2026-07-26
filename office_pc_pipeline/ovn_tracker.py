# -*- coding: utf-8 -*-
"""
오버나잇(OVN) 포지션 추적 — Excel 누적합 방식.

기존 positions.json 체인 의존 구조를 대체한다.
어느 날짜를 재실행해도 Excel 전체내역만으로 결정론적으로 OVN을 계산하므로,
하루가 틀어져도 이후 날짜가 오염되지 않는다.

OVN(당일 진입) = 기준일 시작 포지션(anchor) + Σ(기준일 ≤ 거래일 < 당일 순매수)
당일 종료 포지션      = OVN(당일 진입) + (당일 순매수)

부호 규약: 양수 = 롱, 음수 = 숏. 순매수 = 매수수량 - 매도수량.

월경계 주의: 전체내역은 선물거래_YYYY-MM.xlsx 로 월별 분리 저장된다.
anchor가 이전 달이면 그 달 파일도 함께 읽어야 하므로, anchor월~target월의
모든 월 파일을 순회한다(단일 월만 읽으면 월초에 OVN이 anchor로 되돌아가는 오염 발생).
"""
import json
import re
from datetime import date
from pathlib import Path

import openpyxl

THIS_DIR = Path(__file__).parent
RESULT_DIR = THIS_DIR / "결과"
ANCHOR_FILE = THIS_DIR / "ovn_anchor.json"

# 전체내역(Sheet1) 컬럼 인덱스 (0-based): 출처0 거래일1 계좌2 종목코드3 매수매도4 수량5 ...
COL_DATE = 1
COL_CODE = 3
COL_SIDE = 4
COL_QTY = 5


def _norm_date(v) -> str:
    """셀 값(datetime 또는 문자열)을 'YYYY-MM-DD'로 정규화."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r"(\d{4})[-/](\d{2})[-/](\d{2})", s)
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else s


def load_anchor():
    """
    기준일 포지션 로드. ovn_anchor.json 예:
        {"date": "2026-07-13", "ktb3": 90, "ktb10": -30}
    'date' 시작 시점(그날 체결 반영 전)의 OVN을 의미한다.
    파일이 없으면 (없음, 0, 0) — 전체 히스토리가 0에서 시작한다고 가정.
    """
    if ANCHOR_FILE.exists():
        try:
            d = json.loads(ANCHOR_FILE.read_text(encoding="utf-8"))
            return _norm_date(d.get("date", "")), int(d.get("ktb3", 0) or 0), int(d.get("ktb10", 0) or 0)
        except Exception:
            pass
    return "", 0, 0


def _month_range(start_ym: str, end_ym: str):
    """'YYYY-MM' 시작~끝(포함) 사이의 모든 월을 순서대로 yield."""
    sy, sm = int(start_ym[:4]), int(start_ym[5:7])
    ey, em = int(end_ym[:4]), int(end_ym[5:7])
    y, m = sy, sm
    while (y, m) <= (ey, em):
        yield f"{y:04d}-{m:02d}"
        m += 1
        if m > 12:
            m = 1
            y += 1


def _iter_master_rows(target_date: str, anchor_date: str):
    """
    전체내역 시트의 (거래일, 종목코드, 매수매도, 수량) 행을 순회.
    anchor월~target월의 모든 월 파일을 읽는다(월경계 오염 방지).
    """
    start_ym = anchor_date[:7] if anchor_date else target_date[:7]
    end_ym = target_date[:7]
    seen_any = False
    for ym in _month_range(start_ym, end_ym):
        xlsx_path = RESULT_DIR / f"선물거래_{ym}.xlsx"
        if not xlsx_path.exists():
            continue
        seen_any = True
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb["전체내역"] if "전체내역" in wb.sheetnames else wb.active
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or len(r) <= COL_QTY or r[COL_CODE] is None:
                continue
            yield _norm_date(r[COL_DATE]), str(r[COL_CODE]), str(r[COL_SIDE] or ""), r[COL_QTY]
        wb.close()
    if not seen_any:
        raise FileNotFoundError(
            f"Excel 파일 없음: 결과/선물거래_{start_ym}.xlsx ~ 선물거래_{end_ym}.xlsx")


def compute_ovn(target_date: str):
    """
    target_date: 'YYYY-MM-DD' (또는 'YYYY/MM/DD')
    반환: dict {ovn_ktb3, ovn_ktb10, end_ktb3, end_ktb10, net_ktb3, net_ktb10, anchor_date}
      - ovn_*: 당일 진입 오버나잇
      - net_*: 당일 순매수(매수-매도)
      - end_*: 당일 종료 포지션
    """
    target_date = _norm_date(target_date)
    anchor_date, ovn3, ovn10 = load_anchor()

    net3_today = net10_today = 0
    for d, code, side, qty in _iter_master_rows(target_date, anchor_date):
        try:
            q = int(qty)
        except (TypeError, ValueError):
            continue
        signed = q if side == "매수" else -q if side == "매도" else 0
        if signed == 0:
            continue
        is_ktb3 = code.startswith("A65")
        is_ktb10 = code.startswith("A67")
        if not (is_ktb3 or is_ktb10):
            continue
        if anchor_date and d < anchor_date:
            continue  # 기준일 이전 체결은 anchor에 이미 반영됨
        if d < target_date:
            if is_ktb3:
                ovn3 += signed
            else:
                ovn10 += signed
        elif d == target_date:
            if is_ktb3:
                net3_today += signed
            else:
                net10_today += signed

    return {
        "anchor_date": anchor_date or None,
        "ovn_ktb3": ovn3,
        "ovn_ktb10": ovn10,
        "net_ktb3": net3_today,
        "net_ktb10": net10_today,
        "end_ktb3": ovn3 + net3_today,
        "end_ktb10": ovn10 + net10_today,
    }
