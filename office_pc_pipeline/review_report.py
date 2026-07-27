# -*- coding: utf-8 -*-
"""
기간 리뷰 리포트 — 거래 구분(전략)별 손익 귀속.

사용:
    python review_report.py 2026/07/20 2026/07/27       # 구간
    python review_report.py --week   2026/07/27          # 해당일 포함 주(월~금)
    python review_report.py --month  2026/07             # 월
    python review_report.py --quarter 2026Q3             # 분기

귀속:
  건별손익  = 체결별 실현손익(Excel 전체내역 '손익') → 체결의 구분으로 귀속.
              (NH는 체결별 손익 제공 → 방향성 스캘프 실현이 여기 잡힘.
               SS는 체결별 손익 없음 → 커브/차익의 손익은 아래 갱신차금 중심으로 봄.)
  갱신차금  = 오버나잇 MTM(브로커 확정) → OVN 구분(기본 ovn_default, 예: 커브)으로 귀속.

구분은 trade_tags.json을 매번 실시간 적용(리포트 시점 설정이 진실원천).
"""
import json
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

import classify as _cls

THIS_DIR = Path(__file__).parent
RESULT_DIR = THIS_DIR / "결과"
SNAP_DIR = THIS_DIR / "ovn_snapshots"

COL_SRC, COL_DATE, COL_CODE, COL_SIDE, COL_QTY, COL_TIME, COL_PNL = 0, 1, 3, 4, 5, 7, 9


def _iso(s):
    s = str(s).strip().replace("/", "-")
    return s[:10]


def _month_files(d_from, d_to):
    y, m = int(d_from[:4]), int(d_from[5:7])
    ey, em = int(d_to[:4]), int(d_to[5:7])
    while (y, m) <= (ey, em):
        yield RESULT_DIR / f"선물거래_{y:04d}-{m:02d}.xlsx"
        m += 1
        if m > 12:
            m, y = 1, y + 1


def _fills(d_from, d_to):
    for xlsx in _month_files(d_from, d_to):
        if not xlsx.exists():
            continue
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb["전체내역"] if "전체내역" in wb.sheetnames else wb.active
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or len(r) <= COL_PNL or r[COL_CODE] is None:
                continue
            d = _iso(r[COL_DATE].strftime("%Y-%m-%d") if hasattr(r[COL_DATE], "strftime") else r[COL_DATE])
            if d_from <= d <= d_to:
                yield {
                    "date": d, "source": str(r[COL_SRC] or ""), "code": str(r[COL_CODE]),
                    "side": str(r[COL_SIDE] or ""), "time": str(r[COL_TIME] or ""),
                    "pnl": int(r[COL_PNL] or 0),
                }
        wb.close()


def _snapshots(d_from, d_to):
    if not SNAP_DIR.exists():
        return
    for p in sorted(SNAP_DIR.glob("*.json")):
        try:
            snap = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        d = _iso(snap.get("date", ""))
        if d_from <= d <= d_to:
            yield snap


def build_report(d_from, d_to):
    cfg = _cls.load_cfg()
    by_cat = defaultdict(lambda: {"건별": 0, "갱신차금": 0})

    # 1) 건별손익: 체결별 구분 귀속
    n_fills = 0
    for f in _fills(d_from, d_to):
        n_fills += 1
        cat = _cls.classify(f, cfg)
        by_cat[cat]["건별"] += f["pnl"]

    # 2) 갱신차금: OVN 구분 귀속 (기본 ovn_default, code로 override 가능)
    n_days = 0
    for snap in _snapshots(d_from, d_to):
        n_days += 1
        ren = snap.get("renewal_pdf") or {}
        calc = snap.get("renewal_calc") or {}
        for leg, code in (("ktb3", "A6500"), ("ktb10", "A6700")):
            v = ren.get(leg) or calc.get(leg) or 0
            if not v:
                continue
            cat = _cls.classify({"date": _iso(snap["date"]), "code": code, "ovn": True}, cfg)
            by_cat[cat]["갱신차금"] += int(v)
    return by_cat, n_fills, n_days


def _fmt(v):
    return f"{v/10000:+,.0f}만"


def main():
    args = sys.argv[1:]
    if not args:
        sys.exit(__doc__)

    if args[0] == "--week":
        d = datetime.strptime(_iso(args[1]), "%Y-%m-%d")
        mon = d - timedelta(days=d.weekday())
        d_from, d_to = mon.strftime("%Y-%m-%d"), (mon + timedelta(days=4)).strftime("%Y-%m-%d")
        label = f"주간 {d_from}~{d_to}"
    elif args[0] == "--month":
        ym = args[1].replace("/", "-")
        y, m = int(ym[:4]), int(ym[5:7])
        d_from = f"{y:04d}-{m:02d}-01"
        d_to = f"{y:04d}-{m:02d}-31"
        label = f"월간 {y}-{m:02d}"
    elif args[0] == "--quarter":
        y = int(args[1][:4]); q = int(args[1][-1])
        m0 = (q - 1) * 3 + 1
        d_from = f"{y:04d}-{m0:02d}-01"
        d_to = f"{y:04d}-{m0+2:02d}-31"
        label = f"분기 {y}Q{q}"
    else:
        d_from, d_to = _iso(args[0]), _iso(args[1] if len(args) > 1 else args[0])
        label = f"{d_from}~{d_to}"

    by_cat, n_fills, n_days = build_report(d_from, d_to)

    print(f"\n{'='*58}")
    print(f"  기간 리뷰: {label}   (체결 {n_fills}건 / OVN {n_days}일)")
    print(f"{'='*58}")
    print(f"  {'구분':<8}{'건별손익':>14}{'갱신차금':>14}{'합계':>14}")
    print(f"  {'-'*50}")
    tot_g = tot_r = 0
    for cat, v in sorted(by_cat.items(), key=lambda kv: -(kv[1]['건별'] + kv[1]['갱신차금'])):
        s = v["건별"] + v["갱신차금"]
        tot_g += v["건별"]; tot_r += v["갱신차금"]
        print(f"  {cat:<8}{_fmt(v['건별']):>14}{_fmt(v['갱신차금']):>14}{_fmt(s):>14}")
    print(f"  {'-'*50}")
    print(f"  {'합계':<8}{_fmt(tot_g):>14}{_fmt(tot_r):>14}{_fmt(tot_g+tot_r):>14}")
    print(f"{'='*58}\n")


if __name__ == "__main__":
    main()
