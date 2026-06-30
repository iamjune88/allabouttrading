# -*- coding: utf-8 -*-
"""
파싱된 선물거래 데이터를 Excel에 저장
- 날짜별 시트 분리
- 출처(NH/SS) 컬럼 포함
"""
import os
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MASTER_SHEET = "전체내역"
MASTER_HEADERS = ["출처", "거래일", "계좌", "종목코드", "매수매도", "수량", "가격", "체결시간", "거래금액", "손익", "수수료", "비고"]


def _norm_code(raw) -> str:
    """NH 'A6569000' / SS 'A6569' 모두 'A6569' 형태로 통일"""
    m = re.match(r"^([A-Z]\d{4})", str(raw))
    return m.group(1) if m else str(raw)


def _unify_fills(parsed: dict) -> list[dict]:
    """NH/SS 체결내역을 공통 스키마(MASTER_HEADERS)로 변환"""
    source = parsed.get("source", "")
    date = parsed.get("date", "")
    account = parsed.get("account", "")
    rows = []
    for item in parsed.get("체결", []):
        if source == "NH선물":
            rows.append({
                "출처": source, "거래일": date, "계좌": account,
                "종목코드": _norm_code(item.get("종목")),
                "매수매도": item.get("BS"),
                "수량": item.get("수량"), "가격": item.get("가격"),
                "체결시간": item.get("체결시간", ""),
                "거래금액": item.get("거래금액"), "손익": item.get("손익"),
                "수수료": item.get("수수료"), "비고": "",
            })
        else:  # SS선물
            note = f"{item.get('조건구분', '')}/{item.get('결제구분', '')} #{item.get('번호', '')}"
            rows.append({
                "출처": source, "거래일": date, "계좌": account,
                "종목코드": _norm_code(item.get("종목")),
                "매수매도": item.get("구분"),
                "수량": item.get("수량"), "가격": item.get("가격"),
                "체결시간": item.get("시간", ""),
                "거래금액": "", "손익": "",
                "수수료": "", "비고": note,
            })
    return rows


def _upsert_master(wb, unified_rows: list[dict]):
    """전체내역(Sheet1) 시트에 누적 — 동일 체결건은 중복 추가하지 않음"""
    if MASTER_SHEET in wb.sheetnames:
        ws = wb[MASTER_SHEET]
    else:
        ws = wb.create_sheet(MASTER_SHEET, 0)
        _style_header(ws, 1, len(MASTER_HEADERS))
        for h, val in enumerate(MASTER_HEADERS, 1):
            ws.cell(row=1, column=h, value=val)

    existing_keys = set()
    for r in range(2, ws.max_row + 1):
        key = tuple(ws.cell(row=r, column=c).value for c in (1, 2, 3, 4, 5, 6, 7, 8))
        existing_keys.add(key)

    row = ws.max_row + 1
    for item in unified_rows:
        vals = [item[h] for h in MASTER_HEADERS]
        key = tuple(vals[i] for i in (0, 1, 2, 3, 4, 5, 6, 7))
        if key in existing_keys:
            continue
        for c, v in enumerate(vals, 1):
            ws.cell(row=row, column=c, value=v)
        _style_data(ws, row, len(vals), item["출처"])
        existing_keys.add(key)
        row += 1

    _auto_width(ws)
    ws.freeze_panes = "A2"

    # 시트 순서 항상 맨 앞으로
    if wb.sheetnames[0] != MASTER_SHEET:
        wb.move_sheet(MASTER_SHEET, offset=-wb.sheetnames.index(MASTER_SHEET))


SAVE_DIR = Path(__file__).parent / "결과"
SAVE_DIR.mkdir(exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="맑은 고딕")
DATA_FONT = Font(name="맑은 고딕", size=10)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

NH_COLOR = PatternFill("solid", fgColor="D6E4F0")
SS_COLOR = PatternFill("solid", fgColor="E2EFDA")


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _style_data(ws, row, ncols, source=""):
    fill = NH_COLOR if source == "NH선물" else SS_COLOR if source == "SS선물" else None
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")
        if fill:
            cell.fill = fill


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def _get_or_create_wb(xlsx_path: Path):
    if xlsx_path.exists():
        return openpyxl.load_workbook(xlsx_path)
    return openpyxl.Workbook()


def save_to_excel(parsed_list: list, trade_date: str = None):
    """
    parsed_list: [parse_nh() 또는 parse_ss() 결과, ...]
    trade_date: "YYYY-MM-DD" (없으면 오늘 날짜)
    """
    if not trade_date:
        trade_date = datetime.today().strftime("%Y-%m-%d")

    yyyy_mm = trade_date[:7]  # "2026-06"
    xlsx_path = SAVE_DIR / f"선물거래_{yyyy_mm}.xlsx"

    wb = _get_or_create_wb(xlsx_path)
    sheet_name = trade_date  # "2026-06-16"

    # ── Sheet1: 전체내역 (날짜별 누적, 통합 스키마) ──
    unified_rows = []
    for parsed in parsed_list:
        unified_rows.extend(_unify_fills(parsed))
    _upsert_master(wb, unified_rows)

    # ── Sheet2부터: 날짜별 거래내역 증적 (기존 양식 유지) ──
    # 기존 시트 삭제 후 재생성 (덮어쓰기)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # 기본 시트 제거
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    row = 1

    for parsed in parsed_list:
        source = parsed.get("source", "")
        date = parsed.get("date", trade_date)
        account = parsed.get("account", "")

        # ── 헤더 블록 ──
        ws.cell(row=row, column=1, value=f"▶ {source}  |  거래일: {date}  |  계좌: {account}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, name="맑은 고딕")
        row += 1

        # ── 1. 체결내역 ──
        체결 = parsed.get("체결", [])
        if 체결:
            ws.cell(row=row, column=1, value="[체결내역]").font = Font(bold=True, color="1F4E79", name="맑은 고딕")
            row += 1

            if source == "NH선물":
                headers = ["출처", "거래일", "계좌", "종목", "현재가", "거래구분", "BS", "수량", "가격", "체결시간", "거래금액", "손익", "수수료"]
                _style_header(ws, row, len(headers))
                for h, val in enumerate(headers, 1):
                    ws.cell(row=row, column=h, value=val)
                row += 1
                for item in 체결:
                    vals = [source, date, account,
                            item.get("종목"), item.get("현재가"), item.get("거래구분"),
                            item.get("BS"), item.get("수량"), item.get("가격"),
                            item.get("체결시간"), item.get("거래금액"), item.get("손익"), item.get("수수료")]
                    for c, v in enumerate(vals, 1):
                        ws.cell(row=row, column=c, value=v)
                    _style_data(ws, row, len(vals), source)
                    row += 1

            else:  # SS선물
                headers = ["출처", "거래일", "계좌", "종목", "구분", "수량", "가격", "조건구분", "결제구분", "시간", "번호"]
                _style_header(ws, row, len(headers))
                for h, val in enumerate(headers, 1):
                    ws.cell(row=row, column=h, value=val)
                row += 1
                for item in 체결:
                    vals = [source, date, account,
                            item.get("종목"), item.get("구분"), item.get("수량"),
                            item.get("가격"), item.get("조건구분"), item.get("결제구분"),
                            item.get("시간"), item.get("번호")]
                    for c, v in enumerate(vals, 1):
                        ws.cell(row=row, column=c, value=v)
                    _style_data(ws, row, len(vals), source)
                    row += 1

        # ── 2. SS선물 거래내역 (집계) ──
        거래 = parsed.get("거래", [])
        if 거래:
            ws.cell(row=row, column=1, value="[거래내역 집계]").font = Font(bold=True, color="375623", name="맑은 고딕")
            row += 1
            headers = ["출처", "거래일", "계좌", "종목", "구분", "수량", "체결가", "결제가", "거래금액", "수수료", "실현손익"]
            _style_header(ws, row, len(headers))
            for h, val in enumerate(headers, 1):
                ws.cell(row=row, column=h, value=val)
            row += 1
            for item in 거래:
                vals = [source, date, account,
                        item.get("종목"), item.get("구분"), item.get("수량"),
                        item.get("체결가"), item.get("결제가"), item.get("거래금액"),
                        item.get("수수료"), item.get("실현손익")]
                for c, v in enumerate(vals, 1):
                    ws.cell(row=row, column=c, value=v)
                _style_data(ws, row, len(vals), source)
                row += 1

        # ── 3. 미결잔고 ──
        미결 = parsed.get("미결", [])
        if 미결:
            ws.cell(row=row, column=1, value="[미결잔고]").font = Font(bold=True, color="7030A0", name="맑은 고딕")
            row += 1
            if source == "NH선물":
                headers = ["출처", "거래일", "계좌", "종목", "현재가", "거래구분", "BS", "잔량"]
                _style_header(ws, row, len(headers))
                for h, val in enumerate(headers, 1):
                    ws.cell(row=row, column=h, value=val)
                row += 1
                for item in 미결:
                    vals = [source, date, account,
                            item.get("종목"), item.get("현재가"), item.get("거래구분"),
                            item.get("BS"), item.get("잔량")]
                    for c, v in enumerate(vals, 1):
                        ws.cell(row=row, column=c, value=v)
                    _style_data(ws, row, len(vals), source)
                    row += 1
            else:
                headers = ["출처", "거래일", "계좌", "종목", "구분", "잔량", "미결잔량", "평균매입가", "현재가", "평가손익"]
                _style_header(ws, row, len(headers))
                for h, val in enumerate(headers, 1):
                    ws.cell(row=row, column=h, value=val)
                row += 1
                for item in 미결:
                    vals = [source, date, account,
                            item.get("종목"), item.get("구분"), item.get("잔량"),
                            item.get("미결잔량"), item.get("평균매입가"), item.get("현재가"),
                            item.get("평가손익")]
                    for c, v in enumerate(vals, 1):
                        ws.cell(row=row, column=c, value=v)
                    _style_data(ws, row, len(vals), source)
                    row += 1

        # ── 4. 요약 ──
        요약 = parsed.get("요약", {})
        if 요약:
            ws.cell(row=row, column=1, value="[요약]").font = Font(bold=True, name="맑은 고딕")
            row += 1
            for k, v in 요약.items():
                ws.cell(row=row, column=1, value=k)
                ws.cell(row=row, column=2, value=v)
                row += 1

        row += 2  # 두 회사 사이 공백

    _auto_width(ws)
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)
    print(f"[저장] {xlsx_path}")
    return str(xlsx_path)
