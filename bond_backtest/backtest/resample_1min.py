"""
1분봉 인포맥스 파일 → 5분봉·일봉 파생 파일 생성 (스트리밍, 메모리 안전).
1분봉 파일은 년도별 시트('26'~'10')로 분리 → 전체 시트 순회.
출력(data/): KTB3_5min.csv, KTB3_daily.csv, KTB10_5min.csv, KTB10_daily.csv
"""
from __future__ import annotations
import sys, csv
from pathlib import Path
from datetime import datetime
from collections import OrderedDict

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

from openpyxl import load_workbook

# 원본 1분봉 파일: 기존 시계열 폴더의 파일을 직접 참조 (173MB, 복사 안 함)
SRC = Path(__file__).parent.parent.parent / "시계열" / "선물차트 1분(2010~).xlsx"
OUT = Path(__file__).parent / "data"
OUT.mkdir(exist_ok=True)

DATA_START_ROW = 4  # row1=메타, row2=시리즈명, row3=헤더, row4~=데이터

# openpyxl 1-based 컬럼. 3년=col1~, 10년=col13~ (일자는 3년 col1 공유)
BLOCKS = {
    "KTB3":  {"date": 1,  "time": 2,  "o": 3,  "h": 4,  "l": 5,  "c": 6,  "v": 7,  "fnet": 10},
    "KTB10": {"date": 1,  "time": 13, "o": 14, "h": 15, "l": 16, "c": 17, "v": 18, "fnet": 21},
}


class OHLCAgg:
    def __init__(self):
        self.bars = OrderedDict()

    def add(self, key, o, h, l, c, v, fnet):
        if key not in self.bars:
            self.bars[key] = [o, h, l, c, v, fnet]
        else:
            b = self.bars[key]
            b[1] = max(b[1], h)
            b[2] = min(b[2], l)
            b[3] = c
            b[4] += v
            b[5] = fnet

    def write_csv(self, path):
        with open(path, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh)
            w.writerow(["datetime", "open", "high", "low", "close", "volume", "foreign_net"])
            for key, b in sorted(self.bars.items()):
                w.writerow([key, b[0], b[1], b[2], b[3], b[4], b[5]])
        return len(self.bars)


def _to_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _parse_dt(date_val, time_val):
    if isinstance(date_val, datetime):
        d = date_val.date()
    else:
        try:
            d = datetime.strptime(str(date_val)[:10], "%Y-%m-%d").date()
        except Exception:
            return None
    hh = mm = ss = 0
    if hasattr(time_val, "hour"):
        hh, mm, ss = time_val.hour, time_val.minute, time_val.second
    else:
        s = str(time_val)
        parts = s.split(":")
        try:
            hh = int(parts[0]); mm = int(parts[1]) if len(parts) > 1 else 0
            ss = int(parts[2]) if len(parts) > 2 else 0
        except Exception:
            return None
    return datetime(d.year, d.month, d.day, hh, mm, ss)


def floor_5min(dt):
    return dt.replace(minute=(dt.minute // 5) * 5, second=0, microsecond=0)


def main():
    print(f"[읽기] {SRC.name} (스트리밍, read-only)")
    if not SRC.exists():
        print("  ⚠ 파일 없음:", SRC); return

    wb = load_workbook(SRC, read_only=True, data_only=True)
    sheet_names = sorted(wb.sheetnames, key=lambda s: int(s))
    print(f"  시트 {len(sheet_names)}개 순회: {sheet_names}")

    aggs = {name: {"5min": OHLCAgg(), "daily": OHLCAgg()} for name in BLOCKS}

    row_i = 0
    kept = 0
    for sname_sheet in sheet_names:
        ws = wb[sname_sheet]
        shared_date = None
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            row_i += 1
            date_cell = row[0] if len(row) > 0 else None
            if date_cell is not None:
                shared_date = date_cell

            for name, cmap in BLOCKS.items():
                di = cmap["date"] - 1
                ti = cmap["time"] - 1
                dval = row[di] if di < len(row) and row[di] is not None else shared_date
                tval = row[ti] if ti < len(row) else None
                cval = _to_float(row[cmap["c"] - 1]) if (cmap["c"] - 1) < len(row) else None
                if dval is None or tval is None or cval is None:
                    continue
                dt = _parse_dt(dval, tval)
                if dt is None:
                    continue
                o = _to_float(row[cmap["o"] - 1]); h = _to_float(row[cmap["h"] - 1])
                l = _to_float(row[cmap["l"] - 1]); v = _to_float(row[cmap["v"] - 1]) or 0.0
                fnet = _to_float(row[cmap["fnet"] - 1]) or 0.0
                if o is None or h is None or l is None or cval == 0:
                    continue

                aggs[name]["5min"].add(floor_5min(dt).strftime("%Y-%m-%d %H:%M"),
                                       o, h, l, cval, v, fnet)
                aggs[name]["daily"].add(dt.strftime("%Y-%m-%d"),
                                        o, h, l, cval, v, fnet)
                kept += 1

            if row_i % 200000 == 0:
                print(f"  ...{row_i:,}행 처리 (시트 {sname_sheet})")

    wb.close()
    print(f"[완료] 총 {row_i:,}행 스캔, 유효 집계 {kept:,}건")

    for name in BLOCKS:
        n5 = aggs[name]["5min"].write_csv(OUT / f"{name}_5min.csv")
        nd = aggs[name]["daily"].write_csv(OUT / f"{name}_daily.csv")
        print(f"  {name}: 5분봉 {n5:,}개 | 일봉 {nd:,}개")


if __name__ == "__main__":
    main()
