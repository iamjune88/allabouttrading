"""
체결내역 Excel 증적 관리 모듈

기능:
- 합산된 체결내역(AggregatedFill)을 'Fills' 시트에 누적 append
- 같은 (거래일자, 계좌번호, 종목, 방향, 시간, 가격) 조합은 중복 적재 방지
  (재실행/재수신 시에도 안전하게 idempotent)
- 'Fills' 시트 데이터를 기반으로 포지션 매칭(진입-청산 페어링)을 수행해
  'Trades' 시트에 라운드트립 단위 거래를 별도 기록
- 파일이 없으면 새로 생성, 있으면 로드 후 append
"""
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILLS_HEADERS = [
    "거래일자", "증권사", "계좌번호", "계좌명", "종목", "방향",
    "수량", "가격", "체결시각", "약정금액", "손익", "수수료",
    "원본체결건수", "적재시각",
]

TRADES_HEADERS = [
    "거래일자", "증권사", "계좌번호", "종목",
    "진입시각", "진입방향", "진입수량", "진입가격",
    "청산시각", "청산수량", "청산가격",
    "보유시간(분)", "실현손익", "수수료합계", "순손익", "결과",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def _style_header(ws, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _autosize(ws, ncols: int, min_width=10, max_width=22):
    for c in range(1, ncols + 1):
        col_letter = get_column_letter(c)
        max_len = min_width
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def init_workbook(path: str):
    wb = Workbook()
    ws_fills = wb.active
    ws_fills.title = "Fills"
    ws_fills.append(FILLS_HEADERS)
    _style_header(ws_fills, len(FILLS_HEADERS))
    ws_fills.freeze_panes = "A2"

    ws_trades = wb.create_sheet("Trades")
    ws_trades.append(TRADES_HEADERS)
    _style_header(ws_trades, len(TRADES_HEADERS))
    ws_trades.freeze_panes = "A2"

    wb.save(path)
    return wb


def load_or_init(path: str):
    if Path(path).exists():
        return load_workbook(path)
    return init_workbook(path)


def _existing_fill_keys(ws_fills) -> set[tuple]:
    keys = set()
    for row in ws_fills.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        # (거래일자, 계좌번호, 종목, 방향, 체결시각, 가격)
        keys.add((str(row[0]), str(row[2]), str(row[4]), str(row[5]), str(row[8]), float(row[7])))
    return keys


def append_fills(path: str, aggregated_fills: list) -> dict:
    """
    aggregated_fills: aggregate.AggregatedFill 객체 리스트
    반환: {"inserted": N, "skipped_dup": N}
    """
    wb = load_or_init(path)
    ws = wb["Fills"]
    existing_keys = _existing_fill_keys(ws)

    inserted, skipped = 0, 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for f in aggregated_fills:
        key = (f.trade_date, f.account_no, f.symbol, f.side, f.fill_time, float(f.price))
        if key in existing_keys:
            skipped += 1
            continue
        ws.append([
            f.trade_date, f.broker, f.account_no, f.account_name, f.symbol, f.side,
            f.qty, f.price, f.fill_time, f.notional, f.pnl, f.fee,
            f.raw_count, now,
        ])
        row_idx = ws.max_row
        for c in range(1, len(FILLS_HEADERS) + 1):
            ws.cell(row=row_idx, column=c).font = BODY_FONT
            ws.cell(row=row_idx, column=c).border = THIN_BORDER
        ws.cell(row=row_idx, column=8).number_format = "0.00"
        for c in (10, 11, 12):
            ws.cell(row=row_idx, column=c).number_format = "#,##0;(#,##0);-"
        existing_keys.add(key)
        inserted += 1

    _autosize(ws, len(FILLS_HEADERS))
    wb.save(path)
    return {"inserted": inserted, "skipped_dup": skipped, "total_in_sheet": ws.max_row - 1}


def _existing_trade_keys(ws_trades) -> set[tuple]:
    keys = set()
    # TRADES_HEADERS 인덱스: 0거래일자 1증권사 2계좌번호 3종목 4진입시각 5진입방향
    #                        6진입수량 7진입가격 8청산시각 9청산수량 10청산가격 ...
    for row in ws_trades.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        # (종목, 진입시각, 진입가격, 청산시각, 청산가격, 수량)
        keys.add((str(row[3]), str(row[4]), float(row[7]), str(row[8]), float(row[10]), int(row[6])))
    return keys


def append_trades(path: str, round_trips: list) -> dict:
    """
    round_trips: fifo_matcher.RoundTrip 객체 리스트
    반환: {"inserted": N, "skipped_dup": N}
    """
    wb = load_or_init(path)
    ws = wb["Trades"]
    existing_keys = _existing_trade_keys(ws)

    inserted, skipped = 0, 0
    for r in round_trips:
        key = (r.symbol, r.entry_time, float(r.entry_price), r.exit_time, float(r.exit_price), int(r.qty))
        if key in existing_keys:
            skipped += 1
            continue
        ws.append([
            r.exit_date, r.broker, r.account_no, r.symbol,
            r.entry_time, r.entry_side, r.qty, r.entry_price,
            r.exit_time, r.qty, r.exit_price,
            round(r.holding_minutes, 1) if r.holding_minutes is not None else None,
            round(r.gross_pnl), round(r.fee_total), round(r.net_pnl), r.result,
        ])
        row_idx = ws.max_row
        for c in range(1, len(TRADES_HEADERS) + 1):
            ws.cell(row=row_idx, column=c).font = BODY_FONT
            ws.cell(row=row_idx, column=c).border = THIN_BORDER
        for c in (8, 11):
            ws.cell(row=row_idx, column=c).number_format = "0.00"
        for c in (13, 14, 15):
            ws.cell(row=row_idx, column=c).number_format = "#,##0;(#,##0);-"
        result_cell = ws.cell(row=row_idx, column=16)
        if r.result == "익절":
            result_cell.font = Font(name="Arial", size=10, color="0B6E0B", bold=True)
        elif r.result == "손절":
            result_cell.font = Font(name="Arial", size=10, color="C00000", bold=True)
        existing_keys.add(key)
        inserted += 1

    _autosize(ws, len(TRADES_HEADERS))
    wb.save(path)
    return {"inserted": inserted, "skipped_dup": skipped, "total_in_sheet": ws.max_row - 1}


if __name__ == "__main__":
    import sys
    sys.path.insert(0, "/home/claude/trading_journal")
    from parsers.nh_futures import parse_nh_futures
    from parsers.samsung_futures import parse_samsung_futures
    from parsers.aggregate import aggregate_fills
    from fifo_matcher import match_fifo

    out_path = "/home/claude/trading_journal/output/trading_journal.xlsx"
    Path("/home/claude/trading_journal/output").mkdir(exist_ok=True)

    nh_raws = parse_nh_futures("/mnt/user-data/uploads/20260618_01A103_304526110001_1004.pdf")
    ss_raws = parse_samsung_futures("/mnt/user-data/uploads/20260616100000000000000000014674.pdf")

    nh_aggs = aggregate_fills(nh_raws)
    ss_aggs = aggregate_fills(ss_raws)

    r1 = append_fills(out_path, nh_aggs)
    print("NH선물 적재 결과:", r1)
    r2 = append_fills(out_path, ss_aggs)
    print("삼성선물 적재 결과:", r2)

    # 재실행 시 중복 스킵 확인
    r3 = append_fills(out_path, nh_aggs)
    print("NH선물 재적재(중복테스트) 결과:", r3)

    # FIFO 매칭 -> Trades 시트 적재
    all_fills = nh_aggs + ss_aggs
    round_trips = match_fifo(all_fills)
    r4 = append_trades(out_path, round_trips)
    print("Trades 적재 결과:", r4)
